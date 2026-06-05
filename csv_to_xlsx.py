import csv
import openpyxl
from openpyxl.styles import PatternFill, Font

csv_file = 'fraud_demo_final.csv'
xlsx_file = 'fraud_demo_final.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Данные"

with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for i, row in enumerate(reader):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    ws.column_dimensions[col].width = 16

wb.save(xlsx_file)
print(f'✅ XLSX создан: {xlsx_file}')
