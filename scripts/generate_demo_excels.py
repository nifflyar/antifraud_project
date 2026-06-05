from __future__ import annotations

import csv
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "demo_data" / "excel"
MANIFEST_PATH = ROOT / "demo_data" / "demo_manifest.csv"
README_PATH = ROOT / "demo_data" / "README.md"

HEADERS = [
    "Номер билета",
    "Операция",
    "Номер заказа",
    "Тип тарифа (Льгота)",
    "Цена",
    "Разные сборы",
    "Дата операции",
    "Номер поезда",
    "Класс обслуживания",
    "Станция отправления",
    "Дата/время отправки",
    "Станция назначения",
    "Ф.И.О. пассажира",
    "№ документа",
    "Номер телефона",
    "ИИН",
    "Пол",
    "Канал продаж",
    "Филиал",
    "Агрегатор",
    "Терминал",
    "Пункт продажи",
    "Пользователь",
    "Перевозчик",
    "Тип расчёта",
]

SCENARIOS = [
    "normal_only",
    "similar_refund_cluster",
    "duplicate_ticket_refund",
    "terminal_burst",
    "fake_identity_cluster",
    "close_departure_cluster",
    "seat_blocking",
    "normal_refunds_control",
    "critical_seat_blocking_combo",
    "critical_fake_identity_refund_combo",
    "critical_operator_burst_combo",
    "critical_duplicate_close_combo",
    "document_identity_collision",
    "rapid_cancel_wave",
    "aggregator_refund_wave",
    "route_amount_sweep",
]

STATIONS = [
    "АСТАНА-1",
    "АЛМАТЫ 2",
    "КАРАГАНД П",
    "ШЫМКЕНТ",
    "АКТОБЕ",
    "АТЫРАУ",
    "КОСТАНАЙ",
    "ПАВЛОДАР",
]
CHANNELS = ["Собственные кассы", "Web", "Mobile", "Туристические агентства"]
AGGREGATORS = ["АО \"ПП\" -- --", "Yandex", "Kaspi Travel", "Direct"]
TERMINALS = ["AST-01-001", "ALM-02-014", "KRG-03-009", "SHM-04-020", "RISK-TERM-777"]
CLASSES = ["2С", "2Д", "3П", "3Л", "1Л"]
TARIFFS = ["ПОЛНЫЙ", "ДЕТСКИЙ", "СТУДЕНТ", "ПЕНСИОНЕР", "ЛЬГОТНЫЙ"]
BRANCHES = ["АСТАНА", "АЛМАТЫ", "КАРАГАНДА", "ШЫМКЕНТ"]
USERS = ["cashier.01", "cashier.02", "web.bot", "risk.operator", "mobile.api"]


def iin(seed: int) -> str:
    return f"{860000000000 + seed:012d}"[-12:]


def ticket(seed: int) -> str:
    return f"7{seed:013d}"[-14:]


def passenger(seed: int) -> dict:
    surnames = ["ИВАНОВ", "СЕРИКОВ", "КИМ", "АХМЕТОВА", "ПЕТРОВА", "ЖУМАБЕКОВ"]
    names = ["АРМАН", "МАРАТ", "ЕЛЕНА", "АННА", "ДАНИЯР", "АЙГУЛЬ"]
    middles = ["СЕРИКОВИЧ", "ИВАНОВИЧ", "ПЕТРОВНА", "МАРАТОВНА", "КАНАТОВИЧ"]
    return {
        "fio": f"{surnames[seed % len(surnames)]}={names[seed % len(names)]}={middles[seed % len(middles)]}",
        "iin": iin(seed),
        "doc": f"ID{seed:08d}",
        "phone": f"+7701{seed % 1_000_000:06d}",
        "gender": "Мужской" if seed % 2 == 0 else "Женский",
    }


EXPECTED_SIGNALS = {
    "normal_only": "только обычные продажи и контрольные единичные возвраты",
    "normal_refunds_control": "обычные одиночные возвраты не должны становиться high risk",
    "similar_refund_cluster": "HIGH: несколько похожих возвратов за день",
    "duplicate_ticket_refund": "HIGH: повторный возврат одного билета/заказа",
    "terminal_burst": "операционный риск: кластер возвратов на одном терминале за час",
    "fake_identity_cluster": "HIGH: подозрительная личность + возвраты",
    "close_departure_cluster": "HIGH: возвраты близко к отправлению в кластере",
    "seat_blocking": "CRITICAL: удержание мест + поздние похожие возвраты",
    "critical_seat_blocking_combo": "CRITICAL: seat-blocking + похожие возвраты + близко к отправлению",
    "critical_fake_identity_refund_combo": "CRITICAL: fake identity + organized refund cluster",
    "critical_operator_burst_combo": "CRITICAL: операторский терминальный burst + fake identity",
    "critical_duplicate_close_combo": "CRITICAL: повторные возвраты одного билета близко к отправлению",
    "document_identity_collision": "identity risk: один документ/ИИН используется с разными ФИО",
    "rapid_cancel_wave": "rapid cancellations: оформление и возврат в течение нескольких минут",
    "aggregator_refund_wave": "агрегаторский всплеск возвратов на одном терминале",
    "route_amount_sweep": "серия похожих сумм по одному маршруту",
}


def base_row(
    seed: int,
    op_type: str,
    op_dt: datetime,
    amount: int,
    person: dict,
    *,
    dep_dt: datetime | None = None,
    ticket_no: str | None = None,
    order_no: str | None = None,
    terminal: str | None = None,
    route_seed: int | None = None,
    train_no: str | None = None,
    channel: str | None = None,
    aggregator: str | None = None,
    sale_user: str | None = None,
    point_of_sale: str | None = None,
    branch: str | None = None,
    settlement_type: str | None = None,
) -> dict:
    route_seed = seed if route_seed is None else route_seed
    dep = STATIONS[route_seed % len(STATIONS)]
    arr = STATIONS[(route_seed + 3) % len(STATIONS)]
    dep_dt = dep_dt or (op_dt + timedelta(days=random.randint(3, 25), hours=random.randint(0, 10)))
    ticket_no = ticket_no or ticket(seed)
    order_no = order_no or f"ORD{seed:010d}"
    return {
        "Номер билета": ticket_no,
        "Операция": op_type,
        "Номер заказа": order_no,
        "Тип тарифа (Льгота)": random.choice(TARIFFS),
        "Цена": amount,
        "Разные сборы": random.choice([0, 120, 250]),
        "Дата операции": op_dt,
        "Номер поезда": train_no or f"{(seed % 90) + 1:03d}ЦА",
        "Класс обслуживания": random.choice(CLASSES),
        "Станция отправления": dep,
        "Дата/время отправки": dep_dt,
        "Станция назначения": arr,
        "Ф.И.О. пассажира": person["fio"],
        "№ документа": person["doc"],
        "Номер телефона": person["phone"],
        "ИИН": person["iin"],
        "Пол": person["gender"],
        "Канал продаж": channel or random.choice(CHANNELS),
        "Филиал": branch or random.choice(BRANCHES),
        "Агрегатор": aggregator or random.choice(AGGREGATORS),
        "Терминал": terminal or random.choice(TERMINALS[:-1]),
        "Пункт продажи": point_of_sale or random.choice(["АСТАНА", "АЛМАТЫ", "КАРАГАНДА", "ONLINE"]),
        "Пользователь": sale_user or random.choice(USERS),
        "Перевозчик": "АО ПАССАЖИРСКИЕ ПЕР-КИ",
        "Тип расчёта": settlement_type or random.choice(["Наличный", "Безналичный", "Карта"]),
    }


def add_normal(rows: list[dict], day: datetime, file_idx: int) -> None:
    for n in range(42):
        seed = file_idx * 10_000 + n
        person = passenger(seed)
        op_dt = day.replace(hour=random.randint(7, 22), minute=random.randint(0, 59))
        rows.append(base_row(seed, "Оформление", op_dt, random.randint(3500, 42000), person))

    # Control group: ordinary isolated refunds, far from departure.
    for n in range(4):
        seed = file_idx * 10_000 + 500 + n
        person = passenger(seed)
        op_dt = day.replace(hour=10 + n, minute=15)
        dep_dt = op_dt + timedelta(days=12 + n)
        rows.append(base_row(seed, "Возврат", op_dt, random.randint(5000, 25000), person, dep_dt=dep_dt))


def add_similar_refund_cluster(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(900_000 + file_idx)
    base_amount = 58400 + file_idx * 13
    for n in range(4):
        seed = file_idx * 10_000 + 1000 + n
        op_dt = day.replace(hour=9, minute=10) + timedelta(minutes=n * 18)
        dep_dt = op_dt + timedelta(days=7)
        rows.append(base_row(seed, "Возврат", op_dt, base_amount + random.randint(-350, 350), person, dep_dt=dep_dt, route_seed=4, terminal="RISK-TERM-777"))


def add_duplicate_ticket_refund(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(910_000 + file_idx)
    same_ticket = ticket(file_idx * 777 + 33)
    same_order = f"DUP{file_idx:08d}"
    for n in range(2):
        seed = file_idx * 10_000 + 2000 + n
        op_dt = day.replace(hour=13, minute=5 + n * 22)
        rows.append(base_row(seed, "Возврат", op_dt, 36750, person, ticket_no=same_ticket, order_no=same_order, terminal="RISK-TERM-777"))


def add_terminal_burst(rows: list[dict], day: datetime, file_idx: int) -> None:
    for n in range(6):
        seed = file_idx * 10_000 + 3000 + n
        person = passenger(920_000 + file_idx * 10 + n)
        op_dt = day.replace(hour=16, minute=3 + n * 7)
        rows.append(base_row(seed, "Возврат", op_dt, 22000 + n * 120, person, terminal="RISK-TERM-777"))


def add_fake_identity_cluster(rows: list[dict], day: datetime, file_idx: int) -> None:
    fake = {"fio": "TEST=UNKNOWN=000000", "iin": "", "doc": "", "phone": "", "gender": "Не указан"}
    for n in range(3):
        seed = file_idx * 10_000 + 4000 + n
        op_dt = day.replace(hour=2, minute=12 + n * 11)
        rows.append(base_row(seed, "Возврат", op_dt, 18000 + n * 80, fake, terminal="RISK-TERM-777"))


def add_close_departure_cluster(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(930_000 + file_idx)
    dep_dt = day.replace(hour=21, minute=30)
    for n in range(4):
        seed = file_idx * 10_000 + 5000 + n
        op_dt = dep_dt - timedelta(minutes=55 - n * 8)
        rows.append(base_row(seed, "Возврат", op_dt, 31200 + n * 100, person, dep_dt=dep_dt, terminal="RISK-TERM-777"))


def add_seat_blocking(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(940_000 + file_idx)
    dep_dt = day.replace(hour=23, minute=10)
    train_no = f"9{file_idx % 10:02d}КР"
    for n in range(8):
        seed = file_idx * 10_000 + 6000 + n
        sale_dt = day.replace(hour=8, minute=10 + n)
        rows.append(base_row(seed, "Оформление", sale_dt, 28600, person, dep_dt=dep_dt, route_seed=2, terminal="RISK-TERM-777", train_no=train_no))
    for n in range(5):
        seed = file_idx * 10_000 + 6100 + n
        refund_dt = dep_dt - timedelta(minutes=50 - n * 6)
        rows.append(base_row(seed, "Возврат", refund_dt, 28600, person, dep_dt=dep_dt, route_seed=2, terminal="RISK-TERM-777", train_no=train_no))


def add_critical_seat_blocking_combo(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(1_100_000 + file_idx)
    dep_dt = day.replace(hour=22, minute=40)
    train_no = f"8{file_idx % 10:02d}КР"
    for n in range(10):
        seed = file_idx * 10_000 + 7000 + n
        sale_dt = day.replace(hour=7, minute=5 + n)
        rows.append(base_row(seed, "Оформление", sale_dt, 43800, person, dep_dt=dep_dt, route_seed=5, terminal="RISK-TERM-777", train_no=train_no, channel="Web", aggregator="Kaspi Travel", sale_user="web.bot"))
    for n in range(6):
        seed = file_idx * 10_000 + 7100 + n
        refund_dt = dep_dt - timedelta(minutes=58 - n * 7)
        rows.append(base_row(seed, "Возврат", refund_dt, 43800 + random.randint(-120, 120), person, dep_dt=dep_dt, route_seed=5, terminal="RISK-TERM-777", train_no=train_no, channel="Web", aggregator="Kaspi Travel", sale_user="web.bot"))


def add_critical_fake_identity_refund_combo(rows: list[dict], day: datetime, file_idx: int) -> None:
    fake = {"fio": "TEST=UNKNOWN=000000", "iin": "", "doc": "", "phone": "", "gender": "Не указан"}
    dep_dt = day.replace(hour=19, minute=20)
    train_no = f"7{file_idx % 10:02d}КР"
    for n in range(2):
        seed = file_idx * 10_000 + 8000 + n
        sale_dt = day.replace(hour=1, minute=15 + n)
        rows.append(base_row(seed, "Оформление", sale_dt, 51200, fake, dep_dt=dep_dt, route_seed=1, terminal="RISK-TERM-777", train_no=train_no, channel="Mobile", aggregator="Direct", sale_user="mobile.api"))
    for n in range(5):
        seed = file_idx * 10_000 + 8100 + n
        refund_dt = day.replace(hour=2, minute=10 + n * 9)
        rows.append(base_row(seed, "Возврат", refund_dt, 51200 + random.randint(-180, 180), fake, dep_dt=dep_dt, route_seed=1, terminal="RISK-TERM-777", train_no=train_no, channel="Mobile", aggregator="Direct", sale_user="mobile.api"))


def add_critical_operator_burst_combo(rows: list[dict], day: datetime, file_idx: int) -> None:
    fake = {"fio": "QWERTY=BOT=RISK", "iin": "", "doc": "", "phone": "", "gender": "Не указан"}
    dep_dt = day.replace(hour=18, minute=0)
    train_no = f"6{file_idx % 10:02d}КР"
    for n in range(5):
        seed = file_idx * 10_000 + 9000 + n
        sale_dt = day.replace(hour=11, minute=20 + n)
        rows.append(base_row(seed, "Оформление", sale_dt, 29900, fake, dep_dt=dep_dt, route_seed=3, terminal="RISK-TERM-777", train_no=train_no, channel="Собственные кассы", aggregator="Direct", sale_user="risk.operator", point_of_sale="КАРАГАНДА"))
    for n in range(7):
        seed = file_idx * 10_000 + 9100 + n
        refund_dt = day.replace(hour=15, minute=2 + n * 6)
        rows.append(base_row(seed, "Возврат", refund_dt, 29900 + random.randint(-90, 90), fake, dep_dt=dep_dt, route_seed=3, terminal="RISK-TERM-777", train_no=train_no, channel="Собственные кассы", aggregator="Direct", sale_user="risk.operator", point_of_sale="КАРАГАНДА"))


def add_critical_duplicate_close_combo(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(1_120_000 + file_idx)
    dep_dt = day.replace(hour=21, minute=45)
    train_no = f"5{file_idx % 10:02d}КР"
    same_ticket = ticket(file_idx * 999 + 777)
    same_order = f"CRDUP{file_idx:08d}"
    for n in range(2):
        seed = file_idx * 10_000 + 10_000 + n
        sale_dt = day.replace(hour=9, minute=40 + n)
        rows.append(base_row(seed, "Оформление", sale_dt, 64100, person, dep_dt=dep_dt, ticket_no=same_ticket, order_no=same_order, route_seed=6, terminal="RISK-TERM-777", train_no=train_no))
    for n in range(5):
        seed = file_idx * 10_000 + 10_100 + n
        refund_dt = dep_dt - timedelta(minutes=45 - n * 5)
        rows.append(base_row(seed, "Возврат", refund_dt, 64100, person, dep_dt=dep_dt, ticket_no=same_ticket, order_no=same_order, route_seed=6, terminal="RISK-TERM-777", train_no=train_no))


def add_document_identity_collision(rows: list[dict], day: datetime, file_idx: int) -> None:
    shared_iin = iin(1_130_000 + file_idx)
    shared_doc = f"IDCOL{file_idx:06d}"
    names = ["ИВАНОВ=АРМАН=СЕРИКОВИЧ", "ПЕТРОВА=АННА=ПЕТРОВНА", "КИМ=ЕЛЕНА=МАРАТОВНА"]
    for n, fio in enumerate(names):
        person = {
            "fio": fio,
            "iin": shared_iin,
            "doc": shared_doc,
            "phone": f"+7701555{file_idx:03d}{n}",
            "gender": "Женский" if n else "Мужской",
        }
        seed = file_idx * 10_000 + 11_000 + n
        op_dt = day.replace(hour=12, minute=10 + n * 12)
        rows.append(base_row(seed, "Возврат", op_dt, 18000 + n * 200, person, terminal="RISK-TERM-777", train_no=f"4{file_idx % 10:02d}ИД"))


def add_rapid_cancel_wave(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(1_140_000 + file_idx)
    train_no = f"3{file_idx % 10:02d}РП"
    for n in range(4):
        seed = file_idx * 10_000 + 12_000 + n
        sale_dt = day.replace(hour=10, minute=5 + n * 8)
        dep_dt = sale_dt + timedelta(days=3)
        ticket_no = ticket(seed)
        order_no = f"FAST{file_idx:06d}{n}"
        rows.append(base_row(seed, "Оформление", sale_dt, 27400, person, dep_dt=dep_dt, ticket_no=ticket_no, order_no=order_no, terminal="RISK-TERM-777", train_no=train_no, sale_user="mobile.api"))
        rows.append(base_row(seed + 100, "Возврат", sale_dt + timedelta(minutes=4), 27400, person, dep_dt=dep_dt, ticket_no=ticket_no, order_no=order_no, terminal="RISK-TERM-777", train_no=train_no, sale_user="mobile.api"))


def add_aggregator_refund_wave(rows: list[dict], day: datetime, file_idx: int) -> None:
    for n in range(9):
        person = passenger(1_150_000 + file_idx * 20 + n)
        seed = file_idx * 10_000 + 13_000 + n
        refund_dt = day.replace(hour=17, minute=1 + n * 5)
        rows.append(base_row(seed, "Возврат", refund_dt, 21000 + random.randint(-250, 250), person, terminal="RISK-TERM-777", train_no=f"2{file_idx % 10:02d}АГ", channel="Web", aggregator="Kaspi Travel", sale_user="web.bot"))


def add_route_amount_sweep(rows: list[dict], day: datetime, file_idx: int) -> None:
    person = passenger(1_160_000 + file_idx)
    for n in range(6):
        seed = file_idx * 10_000 + 14_000 + n
        refund_dt = day.replace(hour=14, minute=8) + timedelta(minutes=n * 11)
        rows.append(base_row(seed, "Возврат", refund_dt, 37700 + random.randint(-160, 160), person, route_seed=7, terminal="RISK-TERM-777", train_no=f"1{file_idx % 10:02d}СМ", channel="Туристические агентства", aggregator="Direct", sale_user="risk.operator"))


SCENARIO_BUILDERS = {
    "normal_only": lambda rows, day, idx: None,
    "normal_refunds_control": lambda rows, day, idx: None,
    "similar_refund_cluster": add_similar_refund_cluster,
    "duplicate_ticket_refund": add_duplicate_ticket_refund,
    "terminal_burst": add_terminal_burst,
    "fake_identity_cluster": add_fake_identity_cluster,
    "close_departure_cluster": add_close_departure_cluster,
    "seat_blocking": add_seat_blocking,
    "critical_seat_blocking_combo": add_critical_seat_blocking_combo,
    "critical_fake_identity_refund_combo": add_critical_fake_identity_refund_combo,
    "critical_operator_burst_combo": add_critical_operator_burst_combo,
    "critical_duplicate_close_combo": add_critical_duplicate_close_combo,
    "document_identity_collision": add_document_identity_collision,
    "rapid_cancel_wave": add_rapid_cancel_wave,
    "aggregator_refund_wave": add_aggregator_refund_wave,
    "route_amount_sweep": add_route_amount_sweep,
}


def write_workbook(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "transactions"
    ws.append(HEADERS)

    for row in rows:
        ws.append([row.get(header) for header in HEADERS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    widths = {
        "A": 18,
        "B": 14,
        "C": 18,
        "D": 20,
        "G": 20,
        "H": 14,
        "K": 20,
        "M": 30,
        "N": 16,
        "O": 18,
        "P": 16,
        "R": 22,
        "T": 20,
        "U": 18,
        "V": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main() -> None:
    random.seed(20260518)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)

    for old_file in OUTPUT_DIR.glob("demo_*.xlsx"):
        old_file.unlink()

    start_day = datetime(2026, 3, 16)
    manifest_rows = []
    total_files = 64
    for idx in range(total_files):
        day = start_day + timedelta(days=idx)
        scenario = SCENARIOS[idx % len(SCENARIOS)]
        rows: list[dict] = []
        add_normal(rows, day, idx)
        SCENARIO_BUILDERS[scenario](rows, day, idx)
        random.shuffle(rows)

        filename = f"demo_{idx + 1:02d}_{day:%Y-%m-%d}_{scenario}.xlsx"
        path = OUTPUT_DIR / filename
        write_workbook(path, rows)
        manifest_rows.append(
            {
                "file": filename,
                "date": day.strftime("%Y-%m-%d"),
                "scenario": scenario,
                "rows": len(rows),
                "expected_signal": EXPECTED_SIGNALS[scenario],
            }
        )

    with MANIFEST_PATH.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["file", "date", "scenario", "rows", "expected_signal"])
        writer.writeheader()
        writer.writerows(manifest_rows)

    README_PATH.write_text(
        "# Demo Excel files\n\n"
        f"Папка `excel/` содержит {total_files} XLSX-файла за разные дни недели и месяца.\n"
        "Каждый файл совместим с текущим parser'ом и содержит обычные операции плюс один контролируемый сценарий.\n\n"
        "Ключевые сценарии: ordinary refunds control, similar refund cluster, duplicate ticket refund, terminal burst, fake identity, close departure cluster, seat blocking, critical combo fraud, document collision, rapid cancellations, aggregator wave.\n\n"
        "Critical-сценарии специально комбинируют 2+ сильных сигнала, чтобы passenger-level скоринг показывал `critical`, а operation-level таблица показывала конкретные причины.\n",
        encoding="utf-8",
    )

    print(f"Generated {len(manifest_rows)} files in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
