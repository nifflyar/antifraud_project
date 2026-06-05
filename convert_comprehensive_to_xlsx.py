#!/usr/bin/env python3
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

csv_file = 'comprehensive_test_data.csv'
xlsx_file = 'comprehensive_test_data.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Данные"

# Read CSV and write to Excel
with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for i, row in enumerate(reader):
        ws.append(row)

        # Format header row
        if i == 0:
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")

# Set column widths
col_widths = {
    'A': 20,  # ФИО
    'B': 18,  # ИИН
    'C': 16,  # Документ
    'D': 16,  # Телефон
    'E': 20,  # Email
    'F': 14,  # Тип операции
    'G': 18,  # Дата операции
    'H': 16,  # Дата отправления
    'I': 10,  # Поезд
    'J': 8,   # Вагон
    'K': 8,   # Место
    'L': 14,  # Сумма
    'M': 12,  # Комиссия
    'N': 14,  # Канал
    'O': 14,  # Агрегатор
    'P': 12,  # Статус
    'Q': 14,  # Назначение
}

for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Freeze header
ws.freeze_panes = "A2"

# Add thin border to all cells
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        if cell.row > 1:  # Data rows
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

wb.save(xlsx_file)
print(f'✅ XLSX создан: {xlsx_file}')
print(f'   Всего строк: {ws.max_row - 1}')
print(f'   Формат: готов для загрузки в систему')
