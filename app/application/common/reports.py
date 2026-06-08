from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.application.common.timezone import to_astana_datetime


class ExcelReportGenerator:
    """Helper class to generate formatted Excel reports."""

    def __init__(self, title: str):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self._safe_sheet_title(title)
        self._header_row: int | None = None
        self._set_header_style()

    @staticmethod
    def _safe_sheet_title(title: str) -> str:
        safe_title = "".join("_" if char in r'[]:*?/\\' else char for char in title).strip()
        return (safe_title or "Отчет")[:31]

    def _set_header_style(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def write_headers(self, headers: list[str], row: int = 1):
        self._header_row = row
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center

    def write_rows(self, data: Iterable[Iterable], start_row: int = 2):
        for row_num, row_data in enumerate(data, start_row):
            for col_num, value in enumerate(row_data, 1):
                cell_value = value
                if isinstance(value, datetime):
                    cell_value = to_astana_datetime(value, naive=True)

                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = self.align_left

                if isinstance(value, datetime):
                    cell.number_format = "DD.MM.YYYY HH:MM"
                elif isinstance(value, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(value, int):
                    cell.number_format = "#,##0"

    def _apply_auto_width(self):
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            self.ws.column_dimensions[column].width = min(max(max_length + 2, 10), 48)

    def _apply_table_basics(self):
        if self._header_row and self.ws.max_row >= self._header_row:
            self.ws.freeze_panes = f"A{self._header_row + 1}"
            end_column = get_column_letter(self.ws.max_column)
            self.ws.auto_filter.ref = f"A{self._header_row}:{end_column}{self.ws.max_row}"
        self._apply_auto_width()

    def get_file_bytes(self) -> bytes:
        self._apply_table_basics()
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()
