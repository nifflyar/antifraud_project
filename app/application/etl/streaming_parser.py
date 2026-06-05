"""Streaming Excel parser for large files.

The regular parser is convenient for small files but returns a full list of
rows. This parser keeps memory bounded by yielding validated RawTransaction
chunks while reusing the same column mapping and coercion rules.
"""
from dataclasses import dataclass
import asyncio
import logging
from threading import Thread
from typing import AsyncGenerator, Optional
import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class ParseProgress:
    """Progress update during parsing."""
    rows_parsed: int
    rows_error: int
    batch_number: int
    estimated_total_rows: Optional[int] = None


class StreamingExcelParser:
    """
    Streams Excel file in chunks to avoid memory bloat.
    Yields parsed batches with progress updates.
    """

    def __init__(self, sheet_name: str | int = 0, chunk_size: int = 50000):
        self._sheet_name = sheet_name
        self._chunk_size = chunk_size
        from app.application.etl.excel_parser import (
            ExcelParser,
            ExcelParseError,
        )

        self.ExcelParser = ExcelParser
        self.ExcelParseError = ExcelParseError

    async def parse_stream(
        self, filepath: str
    ) -> AsyncGenerator[tuple[list, list[str], ParseProgress], None]:
        """
        Stream parse Excel file, yielding chunks of (transactions, errors, progress).

        XLSX parsing is CPU/XML-heavy and blocks the interpreter for noticeable
        stretches on 100MB+ workbooks. The parser therefore runs in a dedicated
        worker thread and feeds bounded chunks back to the async ETL loop, while
        database writes stay on the normal SQLAlchemy event loop.
        """
        loop = asyncio.get_running_loop()
        queue: asyncio.Queue[tuple[list, list[str], ParseProgress] | BaseException | object] = asyncio.Queue(maxsize=2)
        sentinel = object()

        def emit(item: tuple[list, list[str], ParseProgress] | BaseException | object) -> None:
            future = asyncio.run_coroutine_threadsafe(queue.put(item), loop)
            future.result()

        def worker() -> None:
            try:
                for item in self._parse_chunks_sync(filepath):
                    emit(item)
            except BaseException as exc:
                emit(exc)
            finally:
                emit(sentinel)

        thread = Thread(target=worker, name="xlsx-parser", daemon=True)
        thread.start()

        while True:
            item = await queue.get()
            if item is sentinel:
                break
            if isinstance(item, BaseException):
                raise item
            yield item

        thread.join(timeout=1)

    def _parse_chunks_sync(
        self, filepath: str
    ):
        """Synchronous XLSX parsing implementation executed in a worker thread."""
        wb = None
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            raise self.ExcelParseError(f"Не удалось открыть файл: {e}") from e

        try:
            ws = (
                wb[self._sheet_name]
                if isinstance(self._sheet_name, str)
                else wb.worksheets[self._sheet_name]
            )

            header_rows = list(ws.iter_rows(max_row=20, values_only=True))
            if not header_rows:
                raise self.ExcelParseError("Файл пустой — нет строк.")
            header_idx, col_map = self.ExcelParser._find_headers(header_rows)
            logger.info(
                "StreamingExcelParser: header row=%s columns=%s",
                header_idx + 1,
                list(col_map.keys()),
            )

            transactions = []
            errors: list[str] = []
            total_parsed = 0
            total_errors = 0
            batch_num = 0

            for row_num, row in enumerate(
                ws.iter_rows(min_row=header_idx + 2, values_only=True),
                start=header_idx + 2,
            ):
                if all(cell is None for cell in row):
                    continue
                if self.ExcelParser._is_non_data_row(row, col_map):
                    continue

                try:
                    transactions.append(self.ExcelParser._parse_row(row, col_map, row_num))
                    total_parsed += 1
                except (ValueError, TypeError) as e:
                    err_msg = f"Строка {row_num}: {e}"
                    errors.append(err_msg)
                    total_errors += 1
                    if total_errors <= 5:
                        logger.warning("StreamingExcelParser parse error: %s", err_msg)

                if len(transactions) + len(errors) >= self._chunk_size:
                    batch_num += 1
                    yield transactions, errors, ParseProgress(
                        rows_parsed=total_parsed,
                        rows_error=total_errors,
                        batch_number=batch_num,
                    )
                    transactions = []
                    errors = []

            if transactions or errors:
                batch_num += 1
                yield transactions, errors, ParseProgress(
                    rows_parsed=total_parsed,
                    rows_error=total_errors,
                    batch_number=batch_num,
                )

            logger.info(
                "StreamingExcelParser complete: parsed=%s errors=%s batches=%s",
                total_parsed,
                total_errors,
                batch_num,
            )
        finally:
            if wb is not None:
                wb.close()
